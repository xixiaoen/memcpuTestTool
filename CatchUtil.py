import subprocess

import time
import openpyxl as op
import os
import Logging
import Constant


def get_cpu_mem(times: int, interval: int, cpu_list=None, mem_list=None):
    """
         抓取指定进程名称的内存、CPU信息
    :param times:
    :param interval:
    :param cpu_list:
    :param mem_list:
    """

    if mem_list is None:
        mem_list = []
    if cpu_list is None:
        cpu_list = []
    Logging.logger.info('>>>------------抓取CPU、MEM数据------------ 开始')
    Logging.logger.info('>>>------------每隔' + str(interval) + '分钟抓取1次，共抓取' + str(times) + '次------------')
    count = times
    i = 0
    cpu_result = dict()
    mem_result = dict()

    while count:
        # 每隔n分钟抓取一次
        time.sleep(60*interval)
        count -= 1
        i += interval
        # 直接把shell命令结果写入文件即可
        Logging.logger.info('>>>------------第' + str(i) + '分钟抓取CPU、MEM数据------------')
        # 获取并写入要查询的进程cpu数据
        cpus = []
        for process in cpu_list:
            # 获取并写入要查询的进程cpu数据
            process_cpu = catch_process_cpu(process)
            cpus.append(process_cpu)
        total_cpu = catch_total_cpu()
        cpus.append(total_cpu)

        cpu_result[i] = cpus

        # 获取并写入要查询的进程mem数据
        mems = []
        for process in mem_list:
            # 获取并写入要查询的进程mem数据
            process_mem = catch_process_mem(process)
            mems.append(process_mem)
        total_mem, free_mem = catch_total_mem()
        mems.extend([free_mem, total_mem])

        mem_result[i] = mems

    Logging.logger.info('>>>------------抓取CPU、MEM数据------------ 完成')
    Logging.logger.info('>>>------------数据写入xls------------ ')
    to_xls(cpu_list, cpu_result, mem_list, mem_result)


def to_xls(cpu_process, cpu_data, mem_process, mem_data):
    cpu_mem_data_dir = Constant.RESULT_PAHT
    cpu_num = Constant.CPU_NUM

    if not os.path.exists(cpu_mem_data_dir):
        os.makedirs(cpu_mem_data_dir)

    # 当前抓取时间
    now_time = time.strftime('%Y-%m-%d_%H%M%S', time.localtime())

    wb = op.Workbook()
    sh = wb.active
    sh.title = "测试结果"
    sh.cell(1, 1, '采集时间点(第N分钟)')
    # 记录平均值列，后续用来合并单元格
    avg_column = []
    # 进程cpu数据起始记录位置
    cpu_start_index = 2
    for i in range(0, len(cpu_process)):
        sh.cell(1, cpu_start_index+i*2, cpu_process[i] + 'CPU占用率')
        sh.cell(1, cpu_start_index+i*2+1, cpu_process[i] + 'CPU占用率平均值')
        avg_column.append(cpu_start_index+i*2+1)
    # 整机cpu数据起始记录位置
    total_cpu_start_index = 2+len(cpu_process)*2
    sh.cell(1, total_cpu_start_index, '整机CPU占用率')
    sh.cell(1, total_cpu_start_index+1, '整机CPU占用率平均值')
    avg_column.append(total_cpu_start_index+1)

    # 进程mem数据起始记录位置
    mem_start_index = 4+len(cpu_process)*2
    for j in range(0, len(mem_process)):
        sh.cell(1, mem_start_index+j*2, mem_process[j] + '内存占用率')
        sh.cell(1, mem_start_index+j*2+1, mem_process[j] + '内存占用率平均值')
        avg_column.append(mem_start_index+j*2+1)
    # 整机mem数据其实记录位置
    total_mem_start_index = 4+len(cpu_process)*2+len(mem_process)*2
    sh.cell(1, total_mem_start_index, '整机内存占用率')
    sh.cell(1, total_mem_start_index+1, '整机内存占用率平均值')
    avg_column.append(total_mem_start_index+1)

    # 写入CPU数据
    row = 2
    for key, value in cpu_data.items():
        Logging.logger.info('%s:%s' % (key, value))
        sh.cell(row, 1, key)
        for k in range(0, len(value)):
            if k == len(value)-1:
                # 整机CPU占用率
                if value[k] != '':
                    data = float(value[k])
                else:
                    data = 0
                data = float('%.4f' % (float(data/100)))
                sh.cell(row, total_cpu_start_index).value = data
                sh.cell(row, total_cpu_start_index).number_format = '0.00%'
            else:
                if value[k] != '':
                    data = float(value[k])/cpu_num
                else:
                    data = 0
                data = float('%.4f' % (float(data/100)))
                sh.cell(row, cpu_start_index+k*2).value = data
                sh.cell(row, cpu_start_index + k * 2).number_format = '0.00%'
        row += 1

    # 写入MEM数据
    row = 2
    for key, value in mem_data.items():
        Logging.logger.info('%s:%s' % (key, value))
        for m in range(0, len(value)-1):
            if m == len(value)-2:
                # 整机内存占用率
                if value[m] != '' and value[len(value)-1] != '':
                    data = 1-float(value[m].replace(',', ''))/float(value[len(value)-1].replace(',', ''))
                else:
                    data = 0
                data = float('%.4f' % (float(data)))
                sh.cell(row, total_mem_start_index).value = data
                sh.cell(row, total_mem_start_index).number_format = '0.00%'
            else:
                if value[m] != '' and value[len(value) - 1] != '':
                    data = float(value[m])/float(value[len(value)-1].replace(',', ''))
                else:
                    data = 0
                data = float('%.4f' % (float(data)))
                sh.cell(row, mem_start_index+m*2).value = data
                sh.cell(row, mem_start_index + m * 2).number_format = '0.00%'
        row += 1

    # 合并单元格并计算平均值
    Logging.logger.info(avg_column)
    for column in avg_column:
        sum = 0
        sh.merge_cells(start_row=2, end_row=row-1, start_column=column, end_column=column)
        for n in range(2, row):
            sum += sh.cell(n, column-1).value
        avg = sum/(row-2)
        avg = float('%.4f' % (float(avg)))
        sh.cell(2, column).value = avg
        sh.cell(2, column).number_format = '0.00%'

    try:
        wb.save(filename=os.path.join(cpu_mem_data_dir, r'cpu_mem_' + now_time + r'.xlsx'))
    except PermissionError:
        pass


def run_shell_cmd(cmds: str):
    cmds = "shell " + cmds
    return run_cmd(cmds)


def run_cmd(cmds: str):
    cmds = "adb " + cmds
    p = subprocess.Popen(cmds, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)

    return p.stdout


def catch_process_cpu(process_name):
    result = run_shell_cmd(f"\"dumpsys cpuinfo | grep {process_name}\"")
    process_cpu = ''
    for line in result.readlines():
        line = line.decode('utf-8')
        Logging.logger.info(process_name + " CPU : " + line.replace('\n', ''))
        if process_name in line:
            process_cpu = line.split('%')[0]
            break
    return process_cpu.strip()


def catch_total_cpu():
    result = run_shell_cmd(f"\"dumpsys cpuinfo | grep 'TOTAL'\"")
    total_cpu = ''
    for line in result.readlines():
        line = line.decode('utf-8')
        Logging.logger.info("TOTAL CPU : " + line.replace('\n', ''))
        if 'TOTAL' in line:
            total_cpu = line.split('%')[0]
            break
    return total_cpu.strip()


def catch_process_mem(process_name):
    result = run_shell_cmd(f"\"dumpsys meminfo {process_name} | grep 'PSS'\"")
    process_mem = ''
    for line in result.readlines():
        line = line.decode('utf-8')
        Logging.logger.info(process_name + " MEM: " + line.replace('\n', ''))
        if 'TOTAL ' in line:
            mem = line.split(':   ')[1]
            process_mem = mem.split('TOTAL ')[0]
            break
    try:
        return process_mem.strip()
    except UnboundLocalError:
        return 0


def catch_total_mem():
    result = run_shell_cmd(f"\"dumpsys meminfo | grep 'RAM'\"")
    total_mem = ''
    free_mem = ''
    for line in result.readlines():
        line = line.decode('utf-8')
        Logging.logger.info("TOTAL MEM: " + line.replace('\n', ''))
        if "Total RAM" in line:
            mem = line.split(': ')[1]
            total_mem = mem.split('K ')[0]
            continue
        if "Free RAM" in line:
            mem = line.split(': ')[1]
            free_mem = mem.split('K ')[0]
            break
    try:
        return total_mem.strip(), free_mem.strip()
    except UnboundLocalError:
        return 0


if __name__ == '__main__':
    # 需要统计的进程列表
    cpuProcessNameList = ["com.boe.team.e3visualprotect", "camerahalserver", "cameraserver"]
    memProcessNameList = ["com.boe.team.e3visualprotect"]
    get_cpu_mem(2, 1, cpuProcessNameList, memProcessNameList)
